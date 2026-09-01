# -*- coding: utf-8 -*-
"""施工曲线2（DB101，替浆期数采）分段分析：排量平台+累计量+碰压时刻。"""
from pathlib import Path
import statistics

import openpyxl

SRC = Path(r"D:\users\desktop\research\控压固井项目\0708\2\HT-002井现场资料\HT1-002五开139.7mm尾管上交\HT1-002施工曲线2.xlsm")
SRC = Path(r"D:\users\desktop\research\控压固井项目\0708\2\HT1-002井现场资料\HT1-002五开139.7mm尾管上交\HT1-002施工曲线2.xlsm")
wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)
ws = wb["原始数据表"]
rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=3, values_only=True))

series = {}
for name, ts, val, *_ in rows:
    if name is None:
        continue
    series.setdefault(name, []).append((ts, val))

press_a = series["DB101_ProcessSensor_PumpPressA_R_FilterValue"]
total = series["DB101_ProcessSensor_TotalAll"]
rate = series["DB1_DAS_PumpRatesAll" if False else "DB101_ProcessSensor_RateAll"]
dens = series["DB101_ProcessSensor_MassFlowDensity_R_ActualValue"]

t0 = min(ts for ts, _ in rate if ts)
t_end = max(ts for ts, _ in rate)
print(f"time span: {t0} -> {t_end}  ({(t_end-t0).total_seconds()/60:.1f} min)")
print(f"points: rate={len(rate)} total={len(total)} dens={len(dens)}")

# 3 分钟桶：排量中位数 + 密度中位数 + 总量增量
import collections
b = collections.defaultdict(lambda: {"r": [], "d": [], "t": []})
for ts, v in rate:
    if ts and v is not None:
        k = int((ts - t0).total_seconds() // 180)
        b[k]["r"].append(v)
for ts, v in dens:
    if ts and v is not None:
        k = int((ts - t0).total_seconds() // 180)
        b[k]["d"].append(v)
for ts, v in total:
    if ts and v is not None:
        k = int((ts - t0).total_seconds() // 180)
        b[k]["t"].append(v)

print(f"{'t_min':>6} {'rate_med':>9} {'dens_med':>9} {'vol_min':>8} {'vol_max':>8}")
last_total = None
for k in sorted(b):
    bk = b[k]
    if not bk["r"]:
        continue
    r_med = statistics.median(bk["r"])
    d_med = statistics.median(bk["d"]) if bk["d"] else float("nan")
    vol_min = min(bk["t"]) if bk["t"] else float("nan")
    vol_max = max(bk["t"] if bk["t"] else [float("nan")])
    print(f"{k*3:>6} {r_med:>9.3f} {d_med:>9.3f} {vol_min:>8.2f} {vol_max:>8.2f}")
