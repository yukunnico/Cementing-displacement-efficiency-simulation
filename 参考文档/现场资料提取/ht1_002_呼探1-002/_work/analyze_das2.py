# -*- coding: utf-8 -*-
"""注灰曲线（DB1，全程数采）排量分段+累计量分析：核对替浆五段排量与总替量。"""
from pathlib import Path
import collections
import statistics

import openpyxl

SRC = Path(r"D:\users\desktop\research\控压固井项目\0708\2\HT1-002井现场资料\HT1-002五开139.7mm尾管上交\HT1-002注灰曲线.xlsm")
wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)
ws = wb["原始数据表"]
rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=3, values_only=True))

series = {}
for name, ts, val, *_ in rows:
    if name is None:
        continue
    series.setdefault(name, []).append((ts, val))

rate = series["DB1_DAS_PumpRatesAll"]
total = series["DB1_DAS_PumpTotalAll"]
t0 = min(ts for ts, _ in total)
t_end = max(ts for ts, _ in total)
print(f"time span: {t0} -> {t_end} ({(t_end-t0).total_seconds()/60:.1f} min)")

b = collections.defaultdict(lambda: {"r": [], "t": []})
for ts, v in rate:
    if ts and v is not None:
        k = int((ts - t0).total_seconds() // 180)
        b[k]["r"].append(v)
for ts, v in total:
    if ts and v is not None:
        k = int((ts - t0).total_seconds() // 180)
        b[k]["t"].append(v)

print(f"{'t_min':>6} {'rate_med':>9} {'vol_start':>10} {'vol_end':>9} {'dvol':>7}")
prev_end = None
for k in sorted(b):
    bk = b[k]
    if not bk["t"]:
        continue
    r_med = statistics.median(bk["r"]) if bk["r"] else 0.0
    vs = min(bk["t"])
    ve = max(bk["t"])
    d = ve - vs
    print(f"{k*3:>6} {r_med:>9.3f} {vs:>10.2f} {ve:>9.2f} {d:>7.2f}")
