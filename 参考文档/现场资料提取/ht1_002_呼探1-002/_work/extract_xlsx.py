# -*- coding: utf-8 -*-
"""提取 hu2 相关全部 Excel（xlsx/xlsm/xls）到 txt：每 sheet 输出非空单元格矩阵。"""
import sys
from pathlib import Path

import openpyxl
import xlrd

OUT_DIR = Path(r"D:\users\desktop\research\控压固井项目\cement model\参考文档\现场资料提取\ht1_002_呼探1-002\_work\xlsx_txt")
OUT_DIR.mkdir(parents=True, exist_ok=True)

BASE = Path(r"D:\users\desktop\research\控压固井项目\0708")
FILES = [
    BASE / "2" / "HT1-002井现场资料" / "HT1-002五开139.7mm尾管上交" / "五开尾管原始记录7-1.xlsx",
    BASE / "2" / "HT1-002井现场资料" / "HT1-002五开139.7mm尾管上交" / "HT1-002施工曲线2.xlsm",
    BASE / "2" / "HT1-002井现场资料" / "HT1-002五开139.7mm尾管上交" / "HT1-002注灰曲线.xlsm",
    BASE / "2" / "HT1-002井现场资料" / "HT1-002井193.7+168.3mm回接上交" / "回接套管数据表.xlsx",
    BASE / "2" / "HT1-002井现场资料" / "HT1-002井193.7+168.3mm回接上交" / "报表 - 呼探1-002井  168.3mm+193.7mm套管扭矩资料.dck.xls",
    BASE / "2" / "HT1-002井现场资料" / "HT1-002井193.7+168.3mm回接上交" / "报表 - 呼探1-002井139.7套管扭矩报表.dck.xls",
    BASE / "1" / "1002" / "10021.xlsx",
    BASE / "1" / "1002" / "10022.xlsx",
    BASE / "2" / "206" / "2061" / "20611" / "206114" / "2061142.xlsx",
    BASE / "2" / "206" / "2061" / "20611" / "206114" / "2061143.xlsx",
    BASE / "2" / "206" / "2061" / "20611" / "206114" / "2061144.xls",
    BASE / "2" / "206" / "2061" / "20611" / "206118" / "四开套管数据.xls",
    BASE / "2" / "206" / "2061" / "20611" / "2061196.xls",
    BASE / "2" / "206" / "2062" / "20626.xls",
    BASE / "2" / "206" / "2063" / "20635.xlsx",
]


def dump_ws(ws, fh, max_rows=400, max_cols=40):
    rows = []
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, max_rows), max_col=min(ws.max_column or 1, max_cols)), 1):
        vals = []
        has_data = False
        for c in row:
            v = c.value
            if v is None:
                vals.append("")
            else:
                has_data = True
                if isinstance(v, float) and v == int(v):
                    v = int(v)
                vals.append(str(v).replace("\n", "\\n"))
        if has_data:
            rows.append(f"R{r}\t" + " | ".join(vals))
    fh.write("\n".join(rows))
    fh.write(f"\n[rows_with_data={len(rows)} max_row={ws.max_row} max_col={ws.max_column}]\n")


def main() -> None:
    for path in FILES:
        tag = path.stem.replace(" ", "_").replace(".", "_")
        out = OUT_DIR / (path.parent.name + "_" + tag + ".txt")
        try:
            if path.suffix.lower() in (".xlsx", ".xlsm"):
                wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
                with out.open("w", encoding="utf-8") as fh:
                    fh.write(f"### {path}\n")
                    for ws in wb.worksheets:
                        fh.write(f"\n#### SHEET: {ws.title} (state={ws.sheet_state})\n")
                        dump_ws(ws, fh)
                wb.close()
            else:  # .xls
                book = xlrd.open_workbook(str(path))
                with out.open("w", encoding="utf-8") as fh:
                    fh.write(f"### {path}\n")
                    for si in range(book.nsheets):
                        sh = book.sheet_by_index(si)
                        fh.write(f"\n#### SHEET: {sh.name} ({sh.nrows}x{sh.ncols})\n")
                        for r in range(min(sh.nrows, 400)):
                            vals = []
                            for c in range(min(sh.ncols, 40)):
                                v = sh.cell_value(r, c)
                                if isinstance(v, float) and v == int(v):
                                    v = int(v)
                                vals.append("" if v == "" else str(v).replace("\n", "\\n"))
                            if any(vals):
                                fh.write(f"R{r+1}\t" + " | ".join(vals) + "\n")
            print(f"[OK] {out.name}  {out.stat().st_size}B")
        except Exception as exc:  # noqa: BLE001
            print(f"[ERR] {path} -> {exc}")


if __name__ == "__main__":
    sys.exit(main())
